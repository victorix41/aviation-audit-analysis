"""Generate management-ready Excel workbooks for aviation audit analytics."""

from datetime import date
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.analytics.corrective_action_engine import (
    generate_corrective_action_analysis,
)
from src.analytics.preventive_action_engine import (
    generate_preventive_action_analysis,
)
from src.ui.data_service import generate_executive_overview
from src.validator import validate_audit_data

REPORT_TITLE = "Aviation MRO Audit Findings Management Workbook"

TITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

SECTION_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="5B9BD5",
)

PASS_FILL = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9",
)

ATTENTION_FILL = PatternFill(
    fill_type="solid",
    fgColor="FCE4D6",
)

WHITE_FONT = Font(
    color="FFFFFF",
    bold=True,
)

TITLE_FONT = Font(
    color="FFFFFF",
    bold=True,
    size=16,
)

SECTION_FONT = Font(
    bold=True,
    size=12,
)

THIN_GREY_BORDER = Border(
    left=Side(
        style="thin",
        color="D9D9D9",
    ),
    right=Side(
        style="thin",
        color="D9D9D9",
    ),
    top=Side(
        style="thin",
        color="D9D9D9",
    ),
    bottom=Side(
        style="thin",
        color="D9D9D9",
    ),
)


def _write_dataframe(
    writer: pd.ExcelWriter,
    dataframe: pd.DataFrame,
    *,
    sheet_name: str,
    startrow: int = 0,
    startcol: int = 0,
    index: bool = False,
) -> None:
    """Write a DataFrame into the workbook."""

    dataframe.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=startrow,
        startcol=startcol,
        index=index,
    )


def _style_title(
    worksheet: Worksheet,
    *,
    title: str,
    end_column: int,
) -> None:
    """Add and format a worksheet title."""

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=end_column,
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
        value=title,
    )

    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[1].height = 28


def _style_section_heading(
    worksheet: Worksheet,
    *,
    row: int,
    title: str,
    end_column: int,
) -> None:
    """Add a section heading across the worksheet."""

    worksheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=end_column,
    )

    cell = worksheet.cell(
        row=row,
        column=1,
        value=title,
    )

    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )


def _style_table_header(
    worksheet: Worksheet,
    *,
    row: int,
    start_column: int,
    end_column: int,
) -> None:
    """Format a table header row."""

    for column in range(
        start_column,
        end_column + 1,
    ):
        cell = worksheet.cell(
            row=row,
            column=column,
        )

        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_GREY_BORDER


def _style_data_region(
    worksheet: Worksheet,
    *,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
) -> None:
    """Apply borders and alignment to a data region."""

    if end_row < start_row:
        return

    for row in worksheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_column,
        max_col=end_column,
    ):
        for cell in row:
            cell.border = THIN_GREY_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def _autofit_columns(
    worksheet: Worksheet,
    *,
    minimum_width: int = 10,
    maximum_width: int = 42,
) -> None:
    """Set readable column widths based on cell contents."""

    for column_cells in worksheet.columns:
        column_index = column_cells[0].column

        if column_index is None:
            continue

        maximum_length = 0

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            maximum_length = max(
                maximum_length,
                len(str(value)),
            )

        adjusted_width = min(
            max(
                maximum_length + 2,
                minimum_width,
            ),
            maximum_width,
        )

        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = adjusted_width


def _add_filter(
    worksheet: Worksheet,
    *,
    header_row: int,
    end_row: int,
    end_column: int,
) -> None:
    """Apply an Excel filter to a table region."""

    if end_row <= header_row:
        return

    worksheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(end_column)}{end_row}"
    )


def _create_executive_summary_sheet(
    writer: pd.ExcelWriter,
    *,
    dataframe: pd.DataFrame,
    source_file_name: str,
    as_of_date: date,
) -> None:
    """Create the Executive Summary worksheet."""

    overview = generate_executive_overview(
        dataframe,
        as_of_date=as_of_date,
    )

    sheet_name = "Executive Summary"

    pd.DataFrame().to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
    )

    worksheet = writer.book[sheet_name]

    _style_title(
        worksheet,
        title=REPORT_TITLE,
        end_column=4,
    )

    worksheet["A3"] = "Source file"
    worksheet["B3"] = source_file_name
    worksheet["A4"] = "As-of date"
    worksheet["B4"] = as_of_date.strftime("%d %B %Y")
    worksheet["A5"] = "Rows analysed"
    worksheet["B5"] = len(dataframe)

    _style_section_heading(
        worksheet,
        row=7,
        title="Finding Profile",
        end_column=4,
    )

    finding_profile = pd.DataFrame(
        {
            "Total Findings": [overview.audit_summary.total_findings],
            "Major": [overview.audit_summary.major_count],
            "Minor": [overview.audit_summary.minor_count],
            "Observations": [overview.audit_summary.observation_count],
        }
    )

    _write_dataframe(
        writer,
        finding_profile,
        sheet_name=sheet_name,
        startrow=7,
    )

    _style_table_header(
        worksheet,
        row=8,
        start_column=1,
        end_column=4,
    )

    _style_data_region(
        worksheet,
        start_row=9,
        end_row=9,
        start_column=1,
        end_column=4,
    )

    _style_section_heading(
        worksheet,
        row=11,
        title="Response Due-Date Position",
        end_column=4,
    )

    due_date_position = pd.DataFrame(
        {
            "Past Due": [overview.audit_summary.past_due_response_count],
            "Due Within 30 Days": [overview.audit_summary.due_within_30_days_count],
            "Future Due": [overview.audit_summary.future_due_count],
            "Missing Due Date": [overview.audit_summary.missing_due_date_count],
        }
    )

    _write_dataframe(
        writer,
        due_date_position,
        sheet_name=sheet_name,
        startrow=11,
    )

    _style_table_header(
        worksheet,
        row=12,
        start_column=1,
        end_column=4,
    )

    _style_data_region(
        worksheet,
        start_row=13,
        end_row=13,
        start_column=1,
        end_column=4,
    )

    _style_section_heading(
        worksheet,
        row=15,
        title="Leading Contributing Categories",
        end_column=4,
    )

    leading_categories = pd.DataFrame(
        [
            {
                "Category": "Human Factor",
                "Leading Result": (
                    overview.human_factor_analysis.top_factor or "Not available"
                ),
                "Frequency": (overview.human_factor_analysis.top_factor_frequency),
                "Percentage": (overview.human_factor_analysis.top_factor_percentage),
            },
            {
                "Category": "Root Cause",
                "Leading Result": (
                    overview.root_cause_analysis.top_root_cause or "Not available"
                ),
                "Frequency": (overview.root_cause_analysis.top_root_cause_frequency),
                "Percentage": (overview.root_cause_analysis.top_root_cause_percentage),
            },
        ]
    )

    _write_dataframe(
        writer,
        leading_categories,
        sheet_name=sheet_name,
        startrow=15,
    )

    _style_table_header(
        worksheet,
        row=16,
        start_column=1,
        end_column=4,
    )

    _style_data_region(
        worksheet,
        start_row=17,
        end_row=18,
        start_column=1,
        end_column=4,
    )

    worksheet["D17"].number_format = "0.00%"
    worksheet["D18"].number_format = "0.00%"

    # The percentages are stored as values from 0 to 100.
    worksheet["D17"] = overview.human_factor_analysis.top_factor_percentage / 100
    worksheet["D18"] = overview.root_cause_analysis.top_root_cause_percentage / 100

    _style_section_heading(
        worksheet,
        row=20,
        title="Management Observations",
        end_column=4,
    )

    current_row = 21

    for observation in overview.executive_insights.observations:
        worksheet.cell(
            row=current_row,
            column=1,
            value=f"• {observation}",
        )

        worksheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=4,
        )

        current_row += 1

    current_row += 1

    _style_section_heading(
        worksheet,
        row=current_row,
        title="Management Considerations",
        end_column=4,
    )

    current_row += 1

    for recommendation in overview.executive_insights.recommendations:
        worksheet.cell(
            row=current_row,
            column=1,
            value=f"• {recommendation}",
        )

        worksheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=4,
        )

        current_row += 1

    worksheet.freeze_panes = "A8"

    _autofit_columns(worksheet)


def _create_analysis_sheet(
    writer: pd.ExcelWriter,
    *,
    sheet_name: str,
    title: str,
    pareto_table: pd.DataFrame,
    monthly_trend: pd.DataFrame,
    quarterly_trend: pd.DataFrame,
    yearly_trend: pd.DataFrame,
) -> None:
    """Create a standard analytics worksheet."""

    pd.DataFrame().to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
    )

    worksheet = writer.book[sheet_name]

    end_column = max(
        4,
        len(monthly_trend.columns),
        len(quarterly_trend.columns),
        len(yearly_trend.columns),
    )

    _style_title(
        worksheet,
        title=title,
        end_column=end_column,
    )

    _style_section_heading(
        worksheet,
        row=3,
        title="Pareto Analysis",
        end_column=end_column,
    )

    pareto_start_row = 3

    _write_dataframe(
        writer,
        pareto_table,
        sheet_name=sheet_name,
        startrow=pareto_start_row,
    )

    pareto_header_row = pareto_start_row + 1
    pareto_end_row = pareto_header_row + len(pareto_table)

    _style_table_header(
        worksheet,
        row=pareto_header_row,
        start_column=1,
        end_column=max(
            len(pareto_table.columns),
            1,
        ),
    )

    _style_data_region(
        worksheet,
        start_row=pareto_header_row + 1,
        end_row=pareto_end_row,
        start_column=1,
        end_column=max(
            len(pareto_table.columns),
            1,
        ),
    )

    current_row = pareto_end_row + 3

    trend_sections = [
        (
            "Monthly Trend",
            monthly_trend,
        ),
        (
            "Quarterly Trend",
            quarterly_trend,
        ),
        (
            "Yearly Trend",
            yearly_trend,
        ),
    ]

    for section_title, trend_table in trend_sections:
        _style_section_heading(
            worksheet,
            row=current_row,
            title=section_title,
            end_column=end_column,
        )

        _write_dataframe(
            writer,
            trend_table,
            sheet_name=sheet_name,
            startrow=current_row,
        )

        header_row = current_row + 1
        end_row = header_row + len(trend_table)

        _style_table_header(
            worksheet,
            row=header_row,
            start_column=1,
            end_column=max(
                len(trend_table.columns),
                1,
            ),
        )

        _style_data_region(
            worksheet,
            start_row=header_row + 1,
            end_row=end_row,
            start_column=1,
            end_column=max(
                len(trend_table.columns),
                1,
            ),
        )

        current_row = end_row + 3

    worksheet.freeze_panes = "A5"

    _autofit_columns(worksheet)


def _build_validation_summary(
    validation_results: dict[str, Any],
) -> pd.DataFrame:
    """Build a validation summary table for Excel."""

    validation_fields = {
        "Duplicate reference numbers": ("duplicate_reference_numbers"),
        "Missing reference numbers": ("missing_reference_numbers"),
        "Missing findings": "missing_findings",
        "Missing due dates": "missing_due_dates",
        "Missing root causes": "missing_root_causes",
        "Missing corrective actions": ("missing_corrective_actions"),
        "Missing preventive actions": ("missing_preventive_actions"),
        "Invalid severity values": ("invalid_severity_values"),
    }

    rows: list[dict[str, Any]] = []

    for label, field_name in validation_fields.items():
        affected = int(
            validation_results.get(
                field_name,
                0,
            )
        )

        rows.append(
            {
                "Validation Check": label,
                "Records Affected": affected,
                "Status": ("Passed" if affected == 0 else "Requires attention"),
            }
        )

    missing_columns = validation_results.get(
        "missing_required_columns",
        [],
    )

    rows.append(
        {
            "Validation Check": ("Missing required columns"),
            "Records Affected": len(missing_columns),
            "Status": ("Passed" if not missing_columns else "Requires attention"),
        }
    )

    return pd.DataFrame(rows)


def _create_data_quality_sheet(
    writer: pd.ExcelWriter,
    *,
    dataframe: pd.DataFrame,
) -> None:
    """Create the Data Quality worksheet."""

    validation_results = validate_audit_data(dataframe)

    validation_summary = _build_validation_summary(validation_results)

    sheet_name = "Data Quality"

    _write_dataframe(
        writer,
        validation_summary,
        sheet_name=sheet_name,
        startrow=2,
    )

    worksheet = writer.book[sheet_name]

    _style_title(
        worksheet,
        title="Data Quality and Validation Summary",
        end_column=5,
    )

    _style_table_header(
        worksheet,
        row=3,
        start_column=1,
        end_column=3,
    )

    _style_data_region(
        worksheet,
        start_row=4,
        end_row=3 + len(validation_summary),
        start_column=1,
        end_column=3,
    )

    for row in range(
        4,
        4 + len(validation_summary),
    ):
        status_cell = worksheet.cell(
            row=row,
            column=3,
        )

        if status_cell.value == "Passed":
            status_cell.fill = PASS_FILL
        else:
            status_cell.fill = ATTENTION_FILL

    missing_summary = pd.DataFrame(
        {
            "Column": dataframe.columns,
            "Missing Values": [
                int(dataframe[column].isna().sum()) for column in dataframe.columns
            ],
        }
    )

    missing_start_row = len(validation_summary) + 6

    _style_section_heading(
        worksheet,
        row=missing_start_row,
        title="Column Missing-Value Summary",
        end_column=5,
    )

    _write_dataframe(
        writer,
        missing_summary,
        sheet_name=sheet_name,
        startrow=missing_start_row,
    )

    _style_table_header(
        worksheet,
        row=missing_start_row + 1,
        start_column=1,
        end_column=2,
    )

    _style_data_region(
        worksheet,
        start_row=missing_start_row + 2,
        end_row=(missing_start_row + 1 + len(missing_summary)),
        start_column=1,
        end_column=2,
    )

    worksheet.freeze_panes = "A4"

    _autofit_columns(worksheet)


def _create_cleaned_data_sheet(
    writer: pd.ExcelWriter,
    *,
    dataframe: pd.DataFrame,
) -> None:
    """Create the cleaned-data worksheet."""

    sheet_name = "Cleaned Audit Data"

    _write_dataframe(
        writer,
        dataframe,
        sheet_name=sheet_name,
    )

    worksheet = writer.book[sheet_name]

    _style_table_header(
        worksheet,
        row=1,
        start_column=1,
        end_column=len(dataframe.columns),
    )

    _style_data_region(
        worksheet,
        start_row=2,
        end_row=len(dataframe) + 1,
        start_column=1,
        end_column=len(dataframe.columns),
    )

    _add_filter(
        worksheet,
        header_row=1,
        end_row=len(dataframe) + 1,
        end_column=len(dataframe.columns),
    )

    worksheet.freeze_panes = "A2"

    for column_index, column_name in enumerate(
        dataframe.columns,
        start=1,
    ):
        if column_name == "response_due_date":
            for row in range(
                2,
                len(dataframe) + 2,
            ):
                worksheet.cell(
                    row=row,
                    column=column_index,
                ).number_format = "yyyy-mm-dd"

    _autofit_columns(
        worksheet,
        maximum_width=38,
    )


def generate_management_excel_report(
    dataframe: pd.DataFrame,
    *,
    source_file_name: str,
    as_of_date: date,
) -> bytes:
    """
    Generate a management-ready Excel analytics workbook.

    Args:
        dataframe:
            Cleaned aviation audit dataset.

        source_file_name:
            Name of the uploaded source file.

        as_of_date:
            Date used for executive due-date classification.

    Returns:
        Excel workbook content as bytes.
    """

    overview = generate_executive_overview(
        dataframe,
        as_of_date=as_of_date,
    )

    corrective_analysis = generate_corrective_action_analysis(dataframe)

    preventive_analysis = generate_preventive_action_analysis(dataframe)

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:
        _create_executive_summary_sheet(
            writer,
            dataframe=dataframe,
            source_file_name=source_file_name,
            as_of_date=as_of_date,
        )

        _create_analysis_sheet(
            writer,
            sheet_name="Severity",
            title="Severity Analysis",
            pareto_table=(overview.severity_analysis.pareto.table),
            monthly_trend=(overview.severity_analysis.monthly_trend),
            quarterly_trend=(overview.severity_analysis.quarterly_trend),
            yearly_trend=(overview.severity_analysis.yearly_trend),
        )

        _create_analysis_sheet(
            writer,
            sheet_name="Human Factors",
            title="Human Factor Analysis",
            pareto_table=(overview.human_factor_analysis.pareto.table),
            monthly_trend=(overview.human_factor_analysis.monthly_trend),
            quarterly_trend=(overview.human_factor_analysis.quarterly_trend),
            yearly_trend=(overview.human_factor_analysis.yearly_trend),
        )

        _create_analysis_sheet(
            writer,
            sheet_name="Root Causes",
            title="Root Cause Analysis",
            pareto_table=(overview.root_cause_analysis.pareto.table),
            monthly_trend=(overview.root_cause_analysis.monthly_wide_trend),
            quarterly_trend=(overview.root_cause_analysis.quarterly_wide_trend),
            yearly_trend=(overview.root_cause_analysis.yearly_wide_trend),
        )

        _create_analysis_sheet(
            writer,
            sheet_name="Corrective Actions",
            title="Corrective Action Analysis",
            pareto_table=(corrective_analysis.pareto.table),
            monthly_trend=(corrective_analysis.monthly_wide_trend),
            quarterly_trend=(corrective_analysis.quarterly_wide_trend),
            yearly_trend=(corrective_analysis.yearly_wide_trend),
        )

        _create_analysis_sheet(
            writer,
            sheet_name="Preventive Actions",
            title="Preventive Action Analysis",
            pareto_table=(preventive_analysis.pareto.table),
            monthly_trend=(preventive_analysis.monthly_wide_trend),
            quarterly_trend=(preventive_analysis.quarterly_wide_trend),
            yearly_trend=(preventive_analysis.yearly_wide_trend),
        )

        _create_data_quality_sheet(
            writer,
            dataframe=dataframe,
        )

        _create_cleaned_data_sheet(
            writer,
            dataframe=dataframe,
        )

        writer.book.properties.title = REPORT_TITLE
        writer.book.properties.subject = "Aviation audit analytics"
        writer.book.properties.creator = "Aviation Audit Analytics"
        writer.book.properties.description = (
            f"Management analytics workbook generated from {source_file_name}."
        )

    return output.getvalue()
