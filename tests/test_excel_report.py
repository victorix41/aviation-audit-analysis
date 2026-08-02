"""Tests for the Excel management workbook generator."""

from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from src.reports.excel_report import (
    REPORT_TITLE,
    generate_management_excel_report,
)

EXPECTED_SHEETS = [
    "Executive Summary",
    "Severity",
    "Human Factors",
    "Root Causes",
    "Corrective Actions",
    "Preventive Actions",
    "Data Quality",
    "Cleaned Audit Data",
]


def create_cleaned_dataframe() -> pd.DataFrame:
    """Create complete cleaned data for workbook tests."""

    return pd.DataFrame(
        {
            "audit_reference_no": [
                "AF-001",
                "AF-002",
                "AF-003",
            ],
            "finding": [
                "Finding one",
                "Finding two",
                "Finding three",
            ],
            "severity_level": [
                "Major",
                "Minor",
                "Observation",
            ],
            "response_due_date": pd.to_datetime(
                [
                    "2026-01-10",
                    "2026-03-10",
                    "2026-04-10",
                ]
            ),
            "immediate_action": [
                "Contain issue",
                "Inspect area",
                "Review record",
            ],
            "root_cause": [
                "Procedure weakness",
                "Training gap",
                "Procedure weakness",
            ],
            "human_factor": [
                "Knowledge gap",
                "Time pressure",
                "Knowledge gap",
            ],
            "corrective_action": [
                "Revise procedure",
                "Provide training",
                "Revise procedure",
            ],
            "preventive_action": [
                "Update audit checklist",
                "Review training programme",
                "Update audit checklist",
            ],
        }
    )


def test_generate_management_excel_report() -> None:
    report_bytes = generate_management_excel_report(
        create_cleaned_dataframe(),
        source_file_name="audit-register.xlsx",
        as_of_date=date(2026, 2, 1),
    )

    assert report_bytes.startswith(b"PK")
    assert len(report_bytes) > 5000


def test_excel_report_contains_expected_sheets() -> None:
    report_bytes = generate_management_excel_report(
        create_cleaned_dataframe(),
        source_file_name="audit-register.xlsx",
        as_of_date=date(2026, 2, 1),
    )

    workbook = load_workbook(
        BytesIO(report_bytes),
        read_only=True,
        data_only=True,
    )

    assert workbook.sheetnames == EXPECTED_SHEETS


def test_excel_report_contains_expected_summary_values() -> None:
    report_bytes = generate_management_excel_report(
        create_cleaned_dataframe(),
        source_file_name="audit-register.xlsx",
        as_of_date=date(2026, 2, 1),
    )

    workbook = load_workbook(
        BytesIO(report_bytes),
        read_only=True,
        data_only=True,
    )

    worksheet = workbook["Executive Summary"]

    assert worksheet["A1"].value == REPORT_TITLE
    assert worksheet["B3"].value == "audit-register.xlsx"
    assert worksheet["B5"].value == 3

    cleaned_sheet = workbook["Cleaned Audit Data"]

    assert cleaned_sheet.max_row == 4
    assert cleaned_sheet.max_column == 9
